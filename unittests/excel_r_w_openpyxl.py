from openpyxl import  Workbook,load_workbook

from numpy import arange
import numpy

PATH = "D:\\PyCharm\\bigdata\\"
wb = Workbook()


###创建Excel 的sheet
wbcreate = wb.create_sheet(index=0,title='oxl_sheet')


##使用numpy的数据类型arange，并自动排列成矩阵8*8
data = arange(1,65).reshape(8,8)

print("shape[0] ",data.shape[0])
print("shape[1] ",data.shape[1])

init_num =0
for c in range(data.shape[0]+init_num):
    print(data.shape[0])
    for r in range(data.shape[1]):
        print(data.shape[1]+init_num)

        ##把array的值写入单元格
        wbcreate.cell(row=r+1,column=c+1).value =data[c,r]

###保存workbook
wb.save(PATH+"excel_r_w_openpyxl.xlsx")



###打开workbook
wb_r = load_workbook(PATH+"excel_r_w_openpyxl.xlsx")
wbread = wb_r.get_active_sheet()

print("active sheet: ",wbread)
print("active sheet  ruturn type: ",type(wbread))
cell = wbread['B4']

lista = []

print(wbread['B1':'B4'])
for cell in  (wbread['B1':'B4']):
    print(cell[0].value)
    lista.append(cell[0].value)
print("lista ",lista)